"""物流报价表解析服务测试。"""

import io
import unittest

from app.services import logistics_quote_parser as parser


def make_csv(text: str) -> bytes:
    return text.encode("utf-8-sig")


CSV_BASIC = """物流报价单（示例）

店家,承运商,线路,时效,首重价,续重价
小鹿家居,中通快运,杭州→上海,1-2天,12元,4.8元/kg
北岸数码,,深圳→北京,2-3天,23,7.5
,,广州→成都,3-4天,,面议
合计,,,,,,
"""


class CsvParsingTests(unittest.TestCase):
    def parse(self, data: bytes, filename: str = "报价.csv"):
        return parser.parse_quote_file(data, filename)

    def test_basic_csv_rows_and_values(self):
        result = self.parse(make_csv(CSV_BASIC))
        self.assertEqual(result["source"]["file_type"], "csv")
        self.assertEqual(result["summary"]["total"], 3)
        self.assertEqual(result["summary"]["valid"], 1)
        self.assertEqual(result["summary"]["review"], 1)
        self.assertEqual(result["summary"]["rejected"], 1)

        first = result["rows"][0]
        self.assertEqual(first["carrier"], "中通快运")
        self.assertEqual(first["seller"], "小鹿家居")
        self.assertEqual(first["route"], "杭州→上海")
        self.assertEqual(first["eta"], "1-2天")
        self.assertEqual(first["first_price"], 12.0)
        self.assertEqual(first["continued_price"], 4.8)
        self.assertEqual(first["review_state"], "valid")
        self.assertIsNone(first["quote"])
        self.assertIn("承运商", first["raw"])
        self.assertEqual(first["raw"]["承运商"], "中通快运")

    def test_row_without_carrier_goes_to_review(self):
        result = self.parse(make_csv(CSV_BASIC))
        second = result["rows"][1]
        self.assertIsNone(second["carrier"])
        self.assertEqual(second["first_price"], 23.0)
        self.assertEqual(second["review_state"], "review")
        self.assertTrue(any("缺少承运商" in issue for issue in second["issues"]))

    def test_unparseable_amount_is_rejected_with_original_text(self):
        result = self.parse(make_csv(CSV_BASIC))
        third = result["rows"][2]
        self.assertEqual(third["review_state"], "rejected")
        self.assertTrue(any("面议" in issue for issue in third["issues"]))

    def test_summary_row_is_skipped_with_warning(self):
        result = self.parse(make_csv(CSV_BASIC))
        self.assertTrue(any("合计" in warning for warning in result["warnings"]))

    def test_gb18030_encoding_supported(self):
        result = self.parse("店家,承运商,首重价,续重价\n店A,中通,12元,4.8\n".encode("gb18030"))
        self.assertEqual(result["summary"]["valid"], 1)
        self.assertEqual(result["rows"][0]["carrier"], "中通")

    def test_tab_delimited_csv_supported(self):
        result = self.parse(make_csv("承运商\t首重价\t续重价\n中通\t12元\t4.8\n"))
        self.assertEqual(result["summary"]["valid"], 1)

    def test_carrier_only_mode_returns_unique_names_without_quote_rows(self):
        result = parser.parse_quote_file(
            make_csv("承运商,首重价\n中通,12\n圆通,13\n中通,14\n"),
            "承运商.csv",
            carrier_only=True,
        )
        self.assertEqual(result["mode"], "carrier_only")
        self.assertEqual(result["summary"], {"total": 2, "valid": 2, "review": 0, "rejected": 0})
        self.assertEqual([carrier["name"] for carrier in result["carriers"]], ["中通", "圆通"])
        self.assertTrue(all(carrier["source"] == "carrier_column" for carrier in result["carriers"]))
        self.assertEqual(result["rows"], [])

    def test_missing_header_fails_cleanly(self):
        with self.assertRaises(ValueError) as ctx:
            self.parse(make_csv("a,b,c\n1,2,3\n"))
        self.assertIn("表头", str(ctx.exception))

    def test_unsupported_extension_rejected(self):
        with self.assertRaises(ValueError):
            self.parse(b"text", "quote.txt")

    def test_old_xls_magic_rejected(self):
        with self.assertRaises(ValueError) as ctx:
            self.parse(b"\xd0\xcf\x11\xe0" + b"\x00" * 16, "报价.xlsx")
        self.assertIn(".xls", str(ctx.exception))

    def test_image_bytes_named_xlsx_rejected(self):
        png = b"\x89PNG\r\n\x1a\n" + b"\x00" * 32
        with self.assertRaises(ValueError) as ctx:
            self.parse(png, "报价.xlsx")
        self.assertIn("图片", str(ctx.exception))

    def test_source_traceability_fields(self):
        result = self.parse(make_csv(CSV_BASIC))
        source = result["source"]
        self.assertEqual(len(source["sha256"]), 64)
        self.assertEqual(source["parser_version"], parser.PARSER_VERSION)
        row = result["rows"][0]
        self.assertEqual(row["sheet"], "CSV")
        self.assertEqual(row["source_row"], 4)
        self.assertEqual(row["id"], "row-CSV-4")


class XlsxParsingTests(unittest.TestCase):
    def setUp(self):
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            self.skipTest("openpyxl 未安装")

    def build_workbook(self) -> bytes:
        import openpyxl

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "华东报价"
        sheet.append(["某店家报价表 2026"])
        sheet.append(["店家", "承运商", "线路", "时效", "首重价", "续重价"])
        sheet.append(["小鹿家居", "中通快运", "杭州→上海", "1-2天", 12, 4.8])
        hidden = workbook.create_sheet("旧数据")
        hidden.sheet_state = "hidden"
        hidden.append(["承运商", "首重价"])
        hidden.append(["废弃", 1])

        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def test_xlsx_rows_parsed_with_sheet_trace(self):
        result = parser.parse_quote_file(self.build_workbook(), "物流价格.xlsx")
        self.assertEqual(result["source"]["file_type"], "xlsx")
        self.assertEqual(result["summary"]["total"], 1)
        self.assertEqual(result["summary"]["valid"], 1)
        row = result["rows"][0]
        self.assertEqual(row["sheet"], "华东报价")
        self.assertEqual(row["source_row"], 3)
        self.assertEqual(row["carrier"], "中通快运")
        self.assertEqual(row["first_price"], 12.0)
        self.assertEqual(row["continued_price"], 4.8)
        self.assertGreater(row["confidence"], 0)

    def test_hidden_sheet_skipped_with_warning(self):
        result = parser.parse_quote_file(self.build_workbook(), "物流价格.xlsx")
        self.assertTrue(any("隐藏" in warning for warning in result["warnings"]))
        self.assertFalse(any(row["sheet"] == "旧数据" for row in result["rows"]))

    def test_ambiguous_header_column_ignored(self):
        import openpyxl

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["店家渠道", "承运商", "首重价", "续重价"])
        sheet.append(["某渠道", "中通", 12, 4.8])
        buffer = io.BytesIO()
        workbook.save(buffer)

        result = parser.parse_quote_file(buffer.getvalue(), "歧义.xlsx")
        row = result["rows"][0]
        self.assertIsNone(row["seller"])
        self.assertTrue(any("不明确" in warning for warning in result["warnings"]))

    def test_corrupted_xlsx_rejected(self):
        with self.assertRaises(ValueError) as ctx:
            parser.parse_quote_file(b"PK\x03\x04not-a-zip", "损坏.xlsx")
        self.assertIn("损坏", str(ctx.exception))


class ThreeCarrierLayoutTests(unittest.TestCase):
    """复刻百世快运/顺心捷达/壹米滴答三张工作表版式的回归样例。"""

    def setUp(self):
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            self.skipTest("openpyxl 未安装")

    def build_workbook(self) -> bytes:
        import openpyxl

        workbook = openpyxl.Workbook()
        first = workbook.active
        first.title = "百世快运"
        first.append([
            "出发省",
            "到达省",
            "首重（KG）",
            "首重价格(元)",
            "0<续重重量≤100kg / 续重价格（元/KG）",
            "100<续重重量≤500kg / 续重价格（元/KG）",
            "续重重量>500kg / 续重价格（元/KG）",
        ])
        first.append(["安徽省", "上海", 30, 42, 1.36, 1.28, 1.18])
        first.append(["北京", "广东省", 30, 45, 1.4, 1.3, 1.2])

        second = workbook.create_sheet("顺心捷达")
        second.append([0, "发件市", "收件省", "收件市", "30KG最低价格", "续重(1KG)价格"])
        second.append(["北京市", "北京市", "安徽省", "合肥市", 50, 1.1])
        second.append(["上海市", "上海市", "浙江省", "杭州市", 55, 1.2])

        third = workbook.create_sheet("壹米滴答")
        third.append(["发件省", "发件市", "收件省", "收件市", "首重(30KG)价格", "续重(1KG)价格"])
        third.append(["安徽省", "合肥市", "上海", "上海", "50.14", "1.49"])

        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def parse(self):
        return parser.parse_quote_file(self.build_workbook(), "物流价格.xlsx")

    def rows_by_sheet(self, result):
        grouped = {}
        for row in result["rows"]:
            grouped.setdefault(row["sheet"], []).append(row)
        return grouped

    def test_carrier_falls_back_to_sheet_name(self):
        result = self.parse()
        grouped = self.rows_by_sheet(result)
        self.assertEqual(result["summary"]["total"], 5)
        for sheet in ("百世快运", "顺心捷达", "壹米滴答"):
            for row in grouped[sheet]:
                self.assertEqual(row["carrier"], sheet)
                self.assertEqual(row["carrier_source"], sheet)
        self.assertTrue(any("承运商来源" in warning for warning in result["warnings"]))

    def test_carrier_only_mode_uses_sheet_names_and_omits_all_rate_rows(self):
        result = parser.parse_quote_file(self.build_workbook(), "物流价格.xlsx", carrier_only=True)
        self.assertEqual([carrier["name"] for carrier in result["carriers"]], ["百世快运", "顺心捷达", "壹米滴答"])
        self.assertTrue(all(carrier["source"] == "sheet_name" for carrier in result["carriers"]))
        self.assertEqual(result["summary"], {"total": 3, "valid": 3, "review": 0, "rejected": 0})
        self.assertEqual(result["rows"], [])

    def test_rate_book_summary_returns_service_chips_without_quote_rows(self):
        result = parser.parse_quote_file(
            self.build_workbook(),
            "物流价格.xlsx",
            rate_book_summary=True,
        )

        self.assertEqual(result["mode"], "rate_book_summary")
        self.assertEqual(result["service_count"], 3)
        self.assertEqual(result["route_count"], 5)
        self.assertEqual(result["book_kind"], "logistics")
        self.assertEqual(result["summary"]["total"], 5)
        self.assertEqual(result["rows"], [])
        self.assertEqual(
            [
                (service["name"], service["rule_type"], service["row_count"], service["route_count"])
                for service in result["services"]
            ],
            [
                ("百世快运", "banded_additional", 2, 2),
                ("顺心捷达", "minimum_then_per_kg", 2, 2),
                ("壹米滴答", "first_additional", 1, 1),
            ],
        )

    def test_tiered_continued_prices_preserved(self):
        result = self.parse()
        row = self.rows_by_sheet(result)["百世快运"][0]
        tiers = row["continued_tiers"]
        self.assertEqual(len(tiers), 3)
        self.assertEqual(
            tiers[0], {"min_exclusive_kg": 0.0, "max_inclusive_kg": 100.0, "price_per_kg": 1.36}
        )
        self.assertEqual(
            tiers[1], {"min_exclusive_kg": 100.0, "max_inclusive_kg": 500.0, "price_per_kg": 1.28}
        )
        self.assertEqual(tiers[2], {"min_exclusive_kg": 500.0, "price_per_kg": 1.18})
        self.assertEqual(row["first_weight_kg"], 30.0)
        self.assertEqual(row["first_price"], 42.0)
        self.assertEqual(row["route"], "安徽省→上海")
        self.assertEqual(row["review_state"], "review")
        self.assertTrue(any("阶梯" in issue for issue in row["issues"]))

    def test_anomalous_numeric_header_becomes_candidate_origin(self):
        result = self.parse()
        row = self.rows_by_sheet(result)["顺心捷达"][0]
        self.assertEqual(row["origin_province"], "北京市")
        self.assertEqual(row["origin_city"], "北京市")
        self.assertEqual(row["destination_province"], "安徽省")
        self.assertEqual(row["first_price"], 50.0)
        self.assertEqual(row["first_weight_kg"], 30.0)
        self.assertEqual(row["continued_unit_kg"], 1.0)
        self.assertEqual(row["route"], "北京市→安徽省合肥市")
        self.assertEqual(row["review_state"], "review")
        self.assertTrue(any("候选" in issue for issue in row["issues"]))
        self.assertTrue(any("候选" in warning for warning in result["warnings"]))

    def test_string_prices_are_numeric_not_weights(self):
        result = self.parse()
        row = self.rows_by_sheet(result)["壹米滴答"][0]
        self.assertEqual(row["first_price"], 50.14)
        self.assertEqual(row["continued_price"], 1.49)
        self.assertEqual(row["first_weight_kg"], 30.0)
        self.assertEqual(row["continued_unit_kg"], 1.0)
        self.assertNotEqual(row["first_weight_kg"], row["first_price"])
        self.assertEqual(row["review_state"], "review")
        self.assertTrue(any("表头推导" in issue for issue in row["issues"]))

    def test_row_cap_warns_explicitly(self):
        import openpyxl
        from unittest import mock

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "承运商"
        sheet.append(["承运商", "首重价", "续重价"])
        for index in range(6):
            sheet.append(["中通", 12, 4.8])
        buffer = io.BytesIO()
        workbook.save(buffer)

        with mock.patch.object(parser, "MAX_ROWS_PER_SHEET", 4):
            result = parser.parse_quote_file(buffer.getvalue(), "截断.xlsx")
        self.assertEqual(result["summary"]["total"], 3)
        self.assertTrue(any("仅解析前 4 行" in warning for warning in result["warnings"]))


class ReportCompatibleRecognitionTests(unittest.TestCase):
    """覆盖参考实现中的通用表头、固定重量档和文件类型约定。"""

    def setUp(self):
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            self.skipTest("openpyxl 未安装")

    def parse_workbook(self, sheet_name, rows, filename="报价.xlsx"):
        import openpyxl

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        for row in rows:
            sheet.append(row)
        buffer = io.BytesIO()
        workbook.save(buffer)
        return parser.parse_quote_file(buffer.getvalue(), filename)

    def test_generic_origin_destination_and_first_additional_rule(self):
        rows = [[f"报价说明 {index}"] for index in range(12)]
        rows.extend(
            [
                ["始发", "目的", "首重(1KG)", "首重价格", "续重(1KG)", "续重价格"],
                ["杭州", "上海", 1, 12, 1, 4.8],
            ]
        )
        result = self.parse_workbook("通用快递", rows)

        self.assertEqual(result["summary"], {"total": 1, "valid": 1, "review": 0, "rejected": 0})
        row = result["rows"][0]
        self.assertEqual(row["carrier"], "通用快递")
        self.assertEqual(row["route"], "杭州→上海")
        self.assertEqual(row["first_weight_kg"], 1.0)
        self.assertEqual(row["continued_unit_kg"], 1.0)
        self.assertEqual(row["rule_type"], "first_additional")
        self.assertEqual(row["book_kind"], "express")
        self.assertEqual(result["service_count"], 1)
        self.assertEqual(result["route_count"], 1)
        self.assertEqual(result["services"][0]["rule_type"], "first_additional")

    def test_fixed_weight_tiers_are_not_treated_as_first_or_continued_prices(self):
        result = self.parse_workbook(
            "固定档报价",
            [
                ["承运商", "起运地", "收件地", "1KG价格", "3KG价格", "5公斤价格"],
                ["中通", "北京", "广州", 12, 19, 27],
            ],
        )

        row = result["rows"][0]
        self.assertEqual(row["route"], "北京→广州")
        self.assertEqual(row["rule_type"], "fixed_tiers")
        self.assertEqual(row["book_kind"], "express")
        self.assertEqual(
            row["fixed_tiers"],
            [
                {"up_to_kg": 1.0, "price": 12.0},
                {"up_to_kg": 3.0, "price": 19.0},
                {"up_to_kg": 5.0, "price": 27.0},
            ],
        )
        self.assertIsNone(row["first_price"])
        self.assertIsNone(row["continued_price"])
        self.assertEqual(row["review_state"], "valid")

    def test_fixed_weight_tiers_with_continued_price_infers_overflow_rule(self):
        result = self.parse_workbook(
            "重量档加续重",
            [
                ["承运商", "发件地", "收件地", "1KG价格", "3KG价格", "续重(1KG)价格"],
                ["德邦", "上海", "深圳", 16, 28, 4.2],
            ],
        )

        row = result["rows"][0]
        self.assertEqual(row["rule_type"], "fixed_tiers_overflow")
        self.assertEqual(row["continued_unit_kg"], 1.0)
        self.assertEqual(row["continued_price"], 4.2)
        self.assertEqual(row["review_state"], "valid")

    def test_minimum_price_rule_and_logistics_book_kind_are_exposed(self):
        result = self.parse_workbook(
            "物流最低价",
            [
                ["承运商", "发件地", "收件地", "30KG最低价格", "续重(1KG)价格"],
                ["顺心", "北京", "合肥", 50, 1.1],
            ],
        )

        row = result["rows"][0]
        self.assertEqual(row["rule_type"], "minimum_then_per_kg")
        self.assertEqual(row["book_kind"], "logistics")
        self.assertEqual(result["book_kind"], "logistics")
        self.assertEqual(row["review_state"], "review")

    def test_xlsm_is_parsed_as_a_supported_excel_workbook(self):
        result = self.parse_workbook(
            "宏工作簿",
            [["承运商", "首重价", "续重价"], ["中通", 12, 4.8]],
            filename="报价.xlsm",
        )
        self.assertEqual(result["source"]["file_type"], "xlsm")
        self.assertEqual(result["summary"]["valid"], 1)

    def test_xls_uses_legacy_parser_when_ole_magic_and_extension_match(self):
        from unittest import mock

        sheets = [("旧版报价", [["承运商", "首重价", "续重价"], ["中通", 12, 4.8]], True)]
        with mock.patch.object(parser, "_iter_xls_rows", return_value=(sheets, [])):
            result = parser.parse_quote_file(parser.OLD_XLS_MAGIC + b"legacy", "报价.xls")

        self.assertEqual(result["source"]["file_type"], "xls")
        self.assertEqual(result["summary"]["valid"], 1)


class AmountParsingTests(unittest.TestCase):
    def test_unit_forms(self):
        cases = {
            "12元": 12.0,
            "4.8元/kg": 4.8,
            "23": 23.0,
            "￥30": 30.0,
            "0.5": 0.5,
        }
        for text, expected in cases.items():
            number, error = parser._parse_amount(text)
            self.assertIsNone(error, text)
            self.assertEqual(number, expected, text)

    def test_ambiguous_forms_not_forced(self):
        for text in ("10-15元", "面议", "12元/票", "12元/单"):
            number, error = parser._parse_amount(text)
            self.assertIsNone(number, text)
            self.assertIsNotNone(error, text)

    def test_negative_amount_flagged(self):
        number, error = parser._parse_amount("-5元")
        self.assertIsNone(number)
        self.assertEqual(error, "金额为负数")


if __name__ == "__main__":
    unittest.main()
