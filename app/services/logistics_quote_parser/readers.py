"""文件类型探测与 CSV/XLSX/XLS 行迭代。"""

from __future__ import annotations

import csv
import io
from typing import Any

from .constants import (
    IMAGE_MAGICS,
    MAX_ROWS_PER_SHEET,
    OLD_XLS_MAGIC,
    WEBP_MAGIC,
    XLSX_MAGIC,
)
from .values import _normalize_text

_SKIP_ROW_KEYWORDS = ("合计", "小计", "总计")


def _iter_csv_rows(data: bytes) -> tuple[list[list[Any]], list[str]]:
    warnings: list[str] = []
    text: str | None = None
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            text = data.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("无法识别文件编码，请导出为 UTF-8 或 GBK 编码的 CSV")
    if "\x00" in text:
        raise ValueError("文件包含二进制内容，不是有效的 CSV")

    sample = text[:4096]
    delimiter = "\t" if sample.count("\t") > sample.count(",") else ","
    try:
        rows = [row for row in csv.reader(io.StringIO(text), delimiter=delimiter)]
    except csv.Error as exc:
        raise ValueError("CSV 文件格式无效，无法解析") from exc
    return rows, warnings


def _iter_xlsx_rows(data: bytes) -> tuple[list[tuple[str, list[list[Any]], bool]], list[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - 部署环境依赖清单中包含 openpyxl
        raise ValueError("服务端缺少 openpyxl 依赖，无法解析 XLSX 文件") from exc

    try:
        workbook = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    except Exception as exc:
        raise ValueError("文件无法作为 XLSX 解析，文件可能已损坏") from exc

    warnings: list[str] = []
    sheets: list[tuple[str, list[list[Any]], bool]] = []
    for name in workbook.sheetnames:
        sheet = workbook[name]
        visible = sheet.sheet_state == "visible"
        rows: list[list[Any]] = []
        for row in sheet.iter_rows(values_only=True):
            rows.append(list(row))
            if len(rows) >= MAX_ROWS_PER_SHEET:
                warnings.append(f"工作表「{name}」超过 {MAX_ROWS_PER_SHEET} 行，仅解析前 {MAX_ROWS_PER_SHEET} 行")
                break
        sheets.append((name, rows, visible))
    workbook.close()
    return sheets, warnings


def _iter_xls_rows(data: bytes) -> tuple[list[tuple[str, list[list[Any]], bool]], list[str]]:
    try:
        import xlrd
    except ImportError as exc:  # pragma: no cover - requirements.txt 声明该依赖
        raise ValueError("服务端缺少 xlrd 依赖，无法解析旧版 .xls 文件") from exc

    try:
        workbook = xlrd.open_workbook(file_contents=data, on_demand=True)
    except Exception as exc:
        raise ValueError("文件无法作为旧版 .xls 解析，文件可能已损坏") from exc

    warnings: list[str] = []
    sheets: list[tuple[str, list[list[Any]], bool]] = []
    visibility = getattr(workbook, "_sheet_visibility", ())
    try:
        for index in range(workbook.nsheets):
            sheet = workbook.sheet_by_index(index)
            visible = not visibility or visibility[index] == 0
            rows = [list(sheet.row_values(row_index)) for row_index in range(min(sheet.nrows, MAX_ROWS_PER_SHEET))]
            if sheet.nrows > MAX_ROWS_PER_SHEET:
                warnings.append(f"工作表「{sheet.name}」超过 {MAX_ROWS_PER_SHEET} 行，仅解析前 {MAX_ROWS_PER_SHEET} 行")
            sheets.append((sheet.name, rows, visible))
    finally:
        workbook.release_resources()
    return sheets, warnings


def _is_skippable_source_row(row: list[Any]) -> bool:
    texts = [_normalize_text(cell) for cell in row]
    if not any(texts):
        return True
    first_cell = next((text for text in texts if text), "")
    return any(keyword in first_cell for keyword in _SKIP_ROW_KEYWORDS)


def _detect_file_type(data: bytes, filename: str) -> str:
    lowered = filename.lower()
    if data[:4] == OLD_XLS_MAGIC:
        if lowered.endswith(".xls"):
            return "xls"
        raise ValueError("文件实际为旧版 .xls 格式，请使用 .xls 扩展名后重新上传")
    if lowered.endswith((".xlsx", ".xlsm")):
        if any(data.startswith(magic) for magic in IMAGE_MAGICS) or (
            data[:4] == WEBP_MAGIC and data[8:12] == b"WEBP"
        ):
            raise ValueError("图片识别尚未接入，请上传 Excel/CSV 格式的报价表")
        if data[:4] != XLSX_MAGIC:
            raise ValueError("文件内容与扩展名不一致，请上传与扩展名相符的报价表")
        return "xlsm" if lowered.endswith(".xlsm") else "xlsx"
    if lowered.endswith(".csv"):
        return "csv"
    if lowered.endswith(".xls"):
        raise ValueError("文件内容与 .xls 扩展名不一致，请上传有效的旧版 Excel 文件")
    raise ValueError("仅支持 .xlsx、.xlsm、.xls 与 .csv 报价表文件")
