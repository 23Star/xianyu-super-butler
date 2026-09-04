"""承运商识别：承运商列/工作表名来源登记与承运商名单摘要模式。"""

from __future__ import annotations

import io
import itertools
from typing import Any

from .constants import HEADER_SCAN_ROWS, MAX_ROWS_PER_SHEET, MAX_SHEETS, MAX_WARNINGS
from .headers import _detect_header, _match_header_field
from .readers import _is_skippable_source_row, _iter_csv_rows
from .values import _cell_text, _normalize_text

_GENERIC_SHEET_NAMES = {"csv", "sheet", "sheet1", "sheet2", "sheet3", "工作表", "工作表1", "worksheet"}


def _detect_carrier_header(rows: list[list[Any]]) -> tuple[int, int, str] | None:
    """在前几行找承运商列；承运商摘要不需要完整的报价字段映射。"""
    for row_index, row in enumerate(rows[:HEADER_SCAN_ROWS]):
        for column, cell in enumerate(row):
            header = _normalize_text(cell)
            if not header:
                continue
            field, conflict = _match_header_field(header)
            if field == "carrier" and conflict is None:
                return row_index, column, _cell_text(cell)
    return None


def _register_carrier(
    registry: dict[str, dict[str, Any]],
    value: Any,
    sheet_name: str,
    source: str,
    carrier_column: str | None = None,
    source_row: int | None = None,
) -> None:
    display_name = str(value or "").strip()
    key = _normalize_text(display_name)
    if not key:
        return

    entry = registry.get(key)
    if entry is None:
        registry[key] = {
            "name": display_name,
            "sheets": [sheet_name],
            "source": source,
            "carrier_column": carrier_column,
            "first_source_row": source_row,
        }
        return

    if sheet_name not in entry["sheets"]:
        entry["sheets"].append(sheet_name)
    if entry["source"] != source:
        entry["source"] = "mixed"
    if entry["carrier_column"] != carrier_column:
        entry["carrier_column"] = None


def _collect_column_carriers(
    registry: dict[str, dict[str, Any]],
    sheet_name: str,
    carrier_column: int,
    carrier_header: str,
    source_rows: Any,
) -> None:
    for source_row, row in source_rows:
        if _is_skippable_source_row(row) or carrier_column >= len(row):
            continue
        _register_carrier(
            registry,
            row[carrier_column],
            sheet_name,
            "carrier_column",
            carrier_header,
            source_row,
        )


def _sheet_name_carrier(sheet_name: str) -> str | None:
    candidate = sheet_name.strip()
    if not candidate or candidate.lower() in _GENERIC_SHEET_NAMES:
        return None
    return candidate


def _carrier_summary_result(
    registry: dict[str, dict[str, Any]],
    mapping: dict[str, str],
    warnings: list[str],
) -> dict[str, Any]:
    carriers = list(registry.values())
    if not carriers:
        raise ValueError("未识别到承运商，请确认工作表名称或承运商列表头")

    services = [
        {
            "name": carrier["name"],
            "sheet_name": carrier["sheets"][0],
            "row_count": 0,
            "route_count": 0,
            "rule_type": "carrier_only",
            "book_kind": None,
            "mapping": {"carrier": carrier["carrier_column"] or "工作表名称"},
        }
        for carrier in carriers
    ]
    return {
        "mode": "carrier_only",
        "mapping": {"matched": mapping, "unmatched": []},
        "summary": {"total": len(carriers), "valid": len(carriers), "review": 0, "rejected": 0},
        "book_kind": None,
        "service_count": len(carriers),
        "route_count": 0,
        "services": services,
        "carriers": carriers,
        "rows": [],
        "warning_count": len(warnings),
        "warnings": warnings[:MAX_WARNINGS],
    }


def _parse_xlsx_carrier_summary(data: bytes) -> tuple[dict[str, dict[str, Any]], dict[str, str], list[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - 部署环境依赖清单中包含 openpyxl
        raise ValueError("服务端缺少 openpyxl 依赖，无法解析 XLSX 文件") from exc

    try:
        workbook = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    except Exception as exc:
        raise ValueError("文件无法作为 XLSX 解析，文件可能已损坏") from exc

    registry: dict[str, dict[str, Any]] = {}
    mapping: dict[str, str] = {}
    warnings: list[str] = []
    sheet_names = workbook.sheetnames
    if len(sheet_names) > MAX_SHEETS:
        warnings.append(f"工作表数量超过 {MAX_SHEETS} 个，仅识别前 {MAX_SHEETS} 个")
        sheet_names = sheet_names[:MAX_SHEETS]

    try:
        for sheet_name in sheet_names:
            sheet = workbook[sheet_name]
            if sheet.sheet_state != "visible":
                warnings.append(f"工作表「{sheet_name}」为隐藏状态，已跳过")
                continue

            row_iterator = sheet.iter_rows(values_only=True)
            header_rows = [list(row) for row in itertools.islice(row_iterator, HEADER_SCAN_ROWS)]
            carrier_header = _detect_carrier_header(header_rows)
            if carrier_header is None:
                header_index, _, _, _ = _detect_header(header_rows)
                candidate = _sheet_name_carrier(sheet_name)
                if candidate and header_index is not None:
                    _register_carrier(registry, candidate, sheet_name, "sheet_name")
                else:
                    warnings.append(f"工作表「{sheet_name}」未识别到承运商列或可识别报价表头")
                continue

            header_index, carrier_column, header_text = carrier_header
            mapping.setdefault("carrier", header_text)
            prefix_rows = enumerate(header_rows[header_index + 1 :], start=header_index + 2)
            remaining_rows = enumerate(row_iterator, start=len(header_rows) + 1)
            row_limit = max(0, MAX_ROWS_PER_SHEET - header_index - 1)
            _collect_column_carriers(
                registry,
                sheet_name,
                carrier_column,
                header_text,
                itertools.islice(itertools.chain(prefix_rows, remaining_rows), row_limit),
            )
            if sheet.max_row > MAX_ROWS_PER_SHEET:
                warnings.append(f"工作表「{sheet_name}」超过 {MAX_ROWS_PER_SHEET} 行，仅识别前 {MAX_ROWS_PER_SHEET} 行")
    finally:
        workbook.close()
    return registry, mapping, warnings


def _parse_xls_carrier_summary(data: bytes) -> tuple[dict[str, dict[str, Any]], dict[str, str], list[str]]:
    try:
        import xlrd
    except ImportError as exc:  # pragma: no cover - requirements.txt 声明该依赖
        raise ValueError("服务端缺少 xlrd 依赖，无法解析旧版 .xls 文件") from exc

    try:
        workbook = xlrd.open_workbook(file_contents=data, on_demand=True)
    except Exception as exc:
        raise ValueError("文件无法作为旧版 .xls 解析，文件可能已损坏") from exc

    registry: dict[str, dict[str, Any]] = {}
    mapping: dict[str, str] = {}
    warnings: list[str] = []
    visibility = getattr(workbook, "_sheet_visibility", ())
    try:
        sheet_count = min(workbook.nsheets, MAX_SHEETS)
        if workbook.nsheets > MAX_SHEETS:
            warnings.append(f"工作表数量超过 {MAX_SHEETS} 个，仅识别前 {MAX_SHEETS} 个")
        for index in range(sheet_count):
            sheet = workbook.sheet_by_index(index)
            visible = not visibility or index >= len(visibility) or visibility[index] == 0
            if not visible:
                warnings.append(f"工作表「{sheet.name}」为隐藏状态，已跳过")
                continue

            header_rows = [list(sheet.row_values(row_index)) for row_index in range(min(sheet.nrows, HEADER_SCAN_ROWS))]
            carrier_header = _detect_carrier_header(header_rows)
            if carrier_header is None:
                header_index, _, _, _ = _detect_header(header_rows)
                candidate = _sheet_name_carrier(sheet.name)
                if candidate and header_index is not None:
                    _register_carrier(registry, candidate, sheet.name, "sheet_name")
                else:
                    warnings.append(f"工作表「{sheet.name}」未识别到承运商列或可识别报价表头")
                continue

            header_index, carrier_column, header_text = carrier_header
            mapping.setdefault("carrier", header_text)
            last_row = min(sheet.nrows, MAX_ROWS_PER_SHEET)
            source_rows = (
                (row_index + 1, list(sheet.row_values(row_index)))
                for row_index in range(header_index + 1, last_row)
            )
            _collect_column_carriers(registry, sheet.name, carrier_column, header_text, source_rows)
            if sheet.nrows > MAX_ROWS_PER_SHEET:
                warnings.append(f"工作表「{sheet.name}」超过 {MAX_ROWS_PER_SHEET} 行，仅识别前 {MAX_ROWS_PER_SHEET} 行")
    finally:
        workbook.release_resources()
    return registry, mapping, warnings


def _parse_carrier_summary(data: bytes, file_type: str) -> dict[str, Any]:
    if file_type == "csv":
        rows, warnings = _iter_csv_rows(data)
        registry: dict[str, dict[str, Any]] = {}
        mapping: dict[str, str] = {}
        carrier_header = _detect_carrier_header(rows)
        if carrier_header is None:
            warnings.append("CSV 未识别到承运商列表头")
        else:
            header_index, carrier_column, header_text = carrier_header
            mapping["carrier"] = header_text
            source_rows = enumerate(rows[header_index + 1 : MAX_ROWS_PER_SHEET], start=header_index + 2)
            _collect_column_carriers(registry, "CSV", carrier_column, header_text, source_rows)
            if len(rows) > MAX_ROWS_PER_SHEET:
                warnings.append(f"CSV 超过 {MAX_ROWS_PER_SHEET} 行，仅识别前 {MAX_ROWS_PER_SHEET} 行")
        return _carrier_summary_result(registry, mapping, warnings)

    if file_type == "xls":
        registry, mapping, warnings = _parse_xls_carrier_summary(data)
    else:
        registry, mapping, warnings = _parse_xlsx_carrier_summary(data)
    return _carrier_summary_result(registry, mapping, warnings)
