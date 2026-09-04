"""报价表摘要模式：为识别结果卡片生成服务级摘要，不保留逐行数据。"""

from __future__ import annotations

import io
import itertools
from typing import Any

from .carriers import _register_carrier, _sheet_name_carrier
from .constants import HEADER_SCAN_ROWS, MAX_ROWS_PER_SHEET, MAX_SHEETS, MAX_WARNINGS
from .headers import _detect_header, _header_implied_weight
from .readers import _is_skippable_source_row, _iter_csv_rows
from .rows import _build_row_result
from .values import _cell_text, _normalize_text


def _new_rate_book_summary_state() -> dict[str, Any]:
    return {
        "warnings": [],
        "mapping": {},
        "unmatched_headers": set(),
        "services": {},
        "carriers": {},
        "valid": 0,
        "review": 0,
        "rejected": 0,
        "total_rows": 0,
        "route_keys": set(),
        "book_kinds": set(),
    }


def _prepare_rate_book_summary_sheet(
    sheet_name: str,
    header_rows: list[list[Any]],
    file_type: str,
    state: dict[str, Any],
) -> dict[str, Any] | None:
    header_index, column_mapping, column_specs, header_warnings = _detect_header(header_rows)
    state["warnings"].extend(header_warnings)
    if header_index is None:
        return None

    header_row = header_rows[header_index]
    sheet_carrier: str | None = None
    if file_type != "csv" and "carrier" not in column_mapping.values():
        sheet_carrier = _sheet_name_carrier(sheet_name)
        if sheet_carrier:
            state["warnings"].append(
                f"工作表「{sheet_name}」未识别到承运商列，已按工作表名作为承运商来源"
            )

    for column, field in column_mapping.items():
        if field == "continued_tier":
            state["mapping"].setdefault("continued_tiers", _cell_text(header_row[column]))
            continue
        if field == "fixed_tier":
            state["mapping"].setdefault("fixed_tiers", _cell_text(header_row[column]))
            continue
        state["mapping"].setdefault(field, _cell_text(header_row[column]))
        if field in ("first_price", "continued_price"):
            header_text = _normalize_text(header_row[column])
            spec = column_specs.setdefault(column, {})
            implied = _header_implied_weight(header_text)
            if implied is not None:
                spec["implied_weight"] = implied
                spec["field"] = field
            if "最低价" in header_text:
                spec["semantics"] = "最低价"
                spec.setdefault("header", _cell_text(header_row[column]))

    for column, cell in enumerate(header_row):
        header = _normalize_text(cell)
        if header and column not in column_mapping:
            state["unmatched_headers"].add(header)

    carrier_column = next(
        (
            _cell_text(header_row[column])
            for column, field in column_mapping.items()
            if field == "carrier"
        ),
        None,
    )
    service_mapping = {
        field: _cell_text(header_row[column])
        for column, field in column_mapping.items()
        if field not in {"continued_tier", "fixed_tier"}
    }
    return {
        "header_index": header_index,
        "column_mapping": column_mapping,
        "column_specs": column_specs,
        "header_cells": {column: _cell_text(header_row[column]) for column in column_mapping},
        "sheet_carrier": sheet_carrier,
        "carrier_column": carrier_column,
        "service_mapping": service_mapping,
    }


def _add_rate_book_summary_row(
    sheet_name: str,
    source_row: int,
    row: list[Any],
    context: dict[str, Any],
    state: dict[str, Any],
) -> None:
    if _is_skippable_source_row(row):
        return

    result = _build_row_result(
        f"summary-{sheet_name}-{source_row}",
        sheet_name,
        source_row,
        context["header_cells"],
        context["column_mapping"],
        row,
        context["column_specs"],
        context["sheet_carrier"],
    )
    state[result["review_state"]] += 1
    state["total_rows"] += 1
    if result["book_kind"]:
        state["book_kinds"].add(result["book_kind"])

    service_name = str(result["carrier"] or context["sheet_carrier"] or sheet_name).strip() or sheet_name
    service_key = (sheet_name, _normalize_text(service_name))
    service = state["services"].get(service_key)
    if service is None:
        source = "carrier_column" if context["carrier_column"] else "sheet_name"
        _register_carrier(
            state["carriers"],
            service_name,
            sheet_name,
            source,
            context["carrier_column"],
            source_row,
        )
        service = {
            "name": service_name,
            "sheet_name": sheet_name,
            "row_count": 0,
            "route_keys": set(),
            "rule_types": set(),
            "book_kinds": set(),
            "mapping": context["service_mapping"],
        }
        state["services"][service_key] = service

    service["row_count"] += 1
    route = result["route"]
    if route:
        service["route_keys"].add(route)
        state["route_keys"].add((service_name, route))
    if result["rule_type"]:
        service["rule_types"].add(result["rule_type"])
    if result["book_kind"]:
        service["book_kinds"].add(result["book_kind"])


def _summarize_rate_book_sheet(
    sheet_name: str,
    header_rows: list[list[Any]],
    remaining_rows: Any,
    total_rows: int,
    file_type: str,
    state: dict[str, Any],
) -> None:
    context = _prepare_rate_book_summary_sheet(sheet_name, header_rows, file_type, state)
    if context is None:
        return

    header_index = context["header_index"]
    prefix_rows = enumerate(header_rows[header_index + 1 :], start=header_index + 2)
    row_limit = max(0, MAX_ROWS_PER_SHEET - header_index - 1)
    source_rows = itertools.islice(itertools.chain(prefix_rows, remaining_rows), row_limit)
    for source_row, row in source_rows:
        _add_rate_book_summary_row(sheet_name, source_row, list(row), context, state)

    if total_rows > MAX_ROWS_PER_SHEET:
        state["warnings"].append(
            f"工作表「{sheet_name}」超过 {MAX_ROWS_PER_SHEET} 行，仅识别前 {MAX_ROWS_PER_SHEET} 行"
        )


def _parse_rate_book_summary(data: bytes, file_type: str) -> dict[str, Any]:
    """生成报价表卡片所需的服务摘要，不保留逐行报价结果。"""
    state = _new_rate_book_summary_state()

    if file_type == "csv":
        rows, csv_warnings = _iter_csv_rows(data)
        state["warnings"].extend(csv_warnings)
        header_rows = rows[:HEADER_SCAN_ROWS]
        _summarize_rate_book_sheet(
            "CSV",
            header_rows,
            enumerate(rows[HEADER_SCAN_ROWS:], start=HEADER_SCAN_ROWS + 1),
            len(rows),
            file_type,
            state,
        )
    elif file_type == "xls":
        try:
            import xlrd
        except ImportError as exc:  # pragma: no cover - requirements.txt 声明该依赖
            raise ValueError("服务端缺少 xlrd 依赖，无法解析旧版 .xls 文件") from exc

        try:
            workbook = xlrd.open_workbook(file_contents=data, on_demand=True)
        except Exception as exc:
            raise ValueError("文件无法作为旧版 .xls 解析，文件可能已损坏") from exc

        visibility = getattr(workbook, "_sheet_visibility", ())
        try:
            sheet_count = min(workbook.nsheets, MAX_SHEETS)
            if workbook.nsheets > MAX_SHEETS:
                state["warnings"].append(f"工作表数量超过 {MAX_SHEETS} 个，仅识别前 {MAX_SHEETS} 个")
            for index in range(sheet_count):
                sheet = workbook.sheet_by_index(index)
                visible = not visibility or index >= len(visibility) or visibility[index] == 0
                if not visible:
                    state["warnings"].append(f"工作表「{sheet.name}」为隐藏状态，已跳过")
                    continue
                header_rows = [
                    list(sheet.row_values(row_index))
                    for row_index in range(min(sheet.nrows, HEADER_SCAN_ROWS))
                ]
                _summarize_rate_book_sheet(
                    sheet.name,
                    header_rows,
                    (
                        (row_index + 1, list(sheet.row_values(row_index)))
                        for row_index in range(HEADER_SCAN_ROWS, sheet.nrows)
                    ),
                    sheet.nrows,
                    file_type,
                    state,
                )
        finally:
            workbook.release_resources()
    else:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - 部署环境依赖清单中包含 openpyxl
            raise ValueError("服务端缺少 openpyxl 依赖，无法解析 XLSX 文件") from exc

        try:
            workbook = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        except Exception as exc:
            raise ValueError("文件无法作为 XLSX 解析，文件可能已损坏") from exc

        sheet_names = workbook.sheetnames[:MAX_SHEETS]
        if len(workbook.sheetnames) > MAX_SHEETS:
            state["warnings"].append(f"工作表数量超过 {MAX_SHEETS} 个，仅识别前 {MAX_SHEETS} 个")
        try:
            for sheet_name in sheet_names:
                sheet = workbook[sheet_name]
                if sheet.sheet_state != "visible":
                    state["warnings"].append(f"工作表「{sheet_name}」为隐藏状态，已跳过")
                    continue
                row_iterator = sheet.iter_rows(values_only=True)
                header_rows = [list(row) for row in itertools.islice(row_iterator, HEADER_SCAN_ROWS)]
                _summarize_rate_book_sheet(
                    sheet_name,
                    header_rows,
                    enumerate(row_iterator, start=len(header_rows) + 1),
                    sheet.max_row,
                    file_type,
                    state,
                )
        finally:
            workbook.close()

    if not state["services"]:
        header_warning = next((warning for warning in state["warnings"] if "表头" in warning), None)
        raise ValueError(header_warning or "未能在文件中解析出报价行，请确认文件是否为报价表")

    if state["unmatched_headers"]:
        sample = "、".join(sorted(state["unmatched_headers"])[:10])
        state["warnings"].append(f"以下列未纳入字段映射：{sample}（可在人工映射中补充）")

    services: list[dict[str, Any]] = []
    for service in state["services"].values():
        rule_types = service["rule_types"]
        book_kinds = service["book_kinds"]
        services.append(
            {
                "name": service["name"],
                "sheet_name": service["sheet_name"],
                "row_count": service["row_count"],
                "route_count": len(service["route_keys"]) or service["row_count"],
                "rule_type": next(iter(rule_types)) if len(rule_types) == 1 else ("mixed" if rule_types else "unknown"),
                "book_kind": next(iter(book_kinds)) if len(book_kinds) == 1 else None,
                "mapping": service["mapping"],
            }
        )

    book_kinds = state["book_kinds"]
    return {
        "mode": "rate_book_summary",
        "mapping": {
            "matched": state["mapping"],
            "unmatched": sorted(state["unmatched_headers"]),
        },
        "summary": {
            "total": state["total_rows"],
            "valid": state["valid"],
            "review": state["review"],
            "rejected": state["rejected"],
        },
        "book_kind": next(iter(book_kinds)) if len(book_kinds) == 1 else None,
        "service_count": len(services),
        "route_count": len(state["route_keys"]) or state["total_rows"],
        "services": services,
        "carriers": list(state["carriers"].values()),
        "rows": [],
        "warning_count": len(state["warnings"]),
        "warnings": state["warnings"][:MAX_WARNINGS],
    }
